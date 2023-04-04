from flask import Blueprint, jsonify, request, g, send_file
from .models import People
from .models import User
from .extentions import db
from functools import wraps
from .utils import generate_token, decode_token
import os
import urllib .parse
from docxtpl import DocxTemplate
from shutil import  rmtree
from  openpyxl import load_workbook

bp = Blueprint('bp', __name__)


@bp.before_app_first_request
def first_request():
    db.create_all()
    try:
        os.mkdir(bp.root_path+"/data")
    except Exception as e:
        print(e)
    try:
        user2 =  User(
                    role = 0,
                    subject_name = "管理员",
                    teacher_name = "老鲁",
                    teacher_info = "我是一个管理员",
                    subject_info = "管理工作好好玩" ,
                    username = "admin",
                    pwd="ldy7842431")
        db.session.add(user2)
        db.session.commit()
    except Exception as e:
        print(e)


@bp.route('/')
def index():
    return jsonify("Hello World!111111111111\n")




@bp.route("/get_admin_infos", methods=['GET'])
def get_admin_infos():
    key_dic = {"教师风采":"fengcai","课时教案":"jiaoan", "社团计划":"jihua", "社团总结":"zongjie", "特色活动方案":"tese", "特色活动图片":"tesetp","精彩瞬间":"shunjian"}
    path = bp.root_path + '/data/'
    response = []
    res = User.query.filter_by(role=1).all()
    for i in res:
        print(i.subject_name, i.teacher_name)
        subject_name = i.subject_name
        teacher_name = i.teacher_name
        obj = {'subject_name':subject_name, 'teacher_name':teacher_name}
        in_path = path + '/' + subject_name
        lst1 = os.listdir(in_path)
        for subdir in lst1:
            in_in_path = in_path + '/' + subdir
            try:
                in_list = os.listdir(in_in_path)
            except Exception as e:
                print(e)
            else:
                obj[key_dic[subdir]] = in_list  
        response.append(obj)
    return jsonify(response)

@bp.route("/get_userlist", methods=['GET'])
def get_userlist():
    res = User.query.all()
    res = list(map(lambda x: {'subject_name': x.subject_name, 
                              'teacher_name': x.teacher_name, 
                              'username': x.username,
                              "role": x.role
                                  }, res))
    return jsonify(res)

@bp.before_request
def judge_login():
    token = request.headers.get('token')
    if not token:
        g.current_user = None
        g.login_message = '尚未登录'
    else:
        res = decode_token(token)
        if not res[0]:
            g.current_user = None
            g.login_message = res[1]
        else:
            g.current_user = res[1]['userid']
            g.login_message = '已登录'


def login_required(func):
    @wraps(func)
    # 获取令牌
    def decorated(*args, **kwargs):
        if g.current_user:
            return func(*args, **kwargs)
        else:
            return jsonify({'message': g.login_message}), 401
    return decorated


@bp.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data['account']
    pwd = data['password']
    res = User.query.filter_by(username=username).first()
    
    if res and res.validate_password(pwd):
        subject_name = res.subject_name
        token = generate_token(res.id)
        role = res.role
        return jsonify({'message': 'success', 'token': token, 'subject_name':subject_name, "role":role}), 200
    else:
        return jsonify({'message':'账号或者密码错误'}), 401




@bp.route('/get_img/<img_name>', methods=['GET'])
def get_img(img_name):
    path = bp.root_path + '/img/'
    return send_file(path+img_name)

@bp.route('/get_tx/<subject_name>/<img_name>', methods=['GET'])
def get_touxiang(subject_name,img_name):
    subject_name = urllib.parse.unquote(subject_name)
    path = bp.root_path +  '/data/'+subject_name+'/教师风采/'
    return send_file(path+img_name)


@bp.route('/get_sj/<subject_name>/<img_name>', methods=['GET'])
def get_shunjian(subject_name,img_name):
    subject_name = urllib.parse.unquote(subject_name)
    path = bp.root_path +  '/data/'+subject_name+'/精彩瞬间/'
    return send_file(path+img_name)
     

@bp.route('/get_ts/<subject_name>/<img_name>', methods=['GET'])
def get_tese(subject_name,img_name):
    subject_name = urllib.parse.unquote(subject_name)
    path = bp.root_path +  '/data/'+subject_name+'/特色活动图片/'
    return send_file(path+img_name)
     


@bp.route('/add_user', methods=['POST'])
def add_sigle_user():
    data = request.get_json()
    if data['role'] == "社团教师":
        data['role'] = 1
    else:
        data['role'] = 0
    res = add_user(data)
    if res == "账户添加成功":
        return jsonify(res)
    else:
        return jsonify({"message": res}), 403
    
    
def add_user(user_info):
    user, subject_name, name, role, pwd = user_info['user'],user_info['subject_name'],user_info['name'], user_info['role'], user_info['pwd'] 
    res = User.query.filter_by(username=user).first()
    if res:
        return "已经存在"
    new_user = User(
                    role = role,
                    subject_name = subject_name,
                    teacher_name = name,
                    teacher_info = "",
                    subject_info = "" ,
                    username = user,
                    pwd=pwd
    )
    db.session.add(new_user)
    
    if role == 1:
        dirname = bp.root_path + '/data/'+new_user.subject_name
        if not os.path.exists(dirname):
            os.mkdir(dirname)
        subdirnames = ["教师风采","课时教案", "社团计划", "社团总结", "特色活动方案", "特色活动图片","精彩瞬间"]
        os.chdir(bp.root_path + '/data/'+new_user.subject_name)
        for sub in subdirnames:
            try:
                os.mkdir(sub)
            except FileExistsError as e:
                print(e)
        db.session.commit()
    
    return "账户添加成功"



@bp.route('/upload_users', methods=['POST'])
def upload_users():
    file = request.files['file']
    file.save(file.filename)
    wb = load_workbook(file.filename)
    st = wb.worksheets[0]
    iter = st.iter_rows(min_row=3)
    succ = []
    fail=[]
    for i in iter:
        infos = list(map(lambda x:x.value, i))
        username, pwd, teacher_name, subject_name, role = infos
        print(infos)
        if(type(pwd)==int):
            pwd = str(pwd)
        try:
             new_user = User(
                    role = role,
                    subject_name = subject_name,
                    teacher_name = teacher_name,
                    teacher_info = "",
                    subject_info = "" ,
                    username = username,
                    pwd=pwd)
             db.session.add(new_user)
             db.session.commit()
        except Exception as  e:
            fail.append(subject_name)
            print(e)
        else:
            dirname = bp.root_path + '/data/'+new_user.subject_name
            if not os.path.exists(dirname):
                os.mkdir(dirname)
            subdirnames = ["教师风采","课时教案", "社团计划", "社团总结", "特色活动方案", "特色活动图片","精彩瞬间"]
            os.chdir(bp.root_path + '/data/'+new_user.subject_name)
            for sub in subdirnames:
                try:
                    os.mkdir(sub)
                except FileExistsError as e:
                    print(e)
            succ.append(subject_name)
            
                
        
    return jsonify({
        "success":succ,
        "fail":fail
    })



@bp.route("/reset_password", methods=['POST'])
def reset_password():
    json = request.get_json()
    username = json['username']
    res = User.query.filter_by(username=username).first()
    res.set_password("88888888")
    db.session.commit()
    return "hahah "



@bp.route("/remove_user", methods=['POST'])
def remove_user():
     json = request.get_json()
     username = json['username']
     res = User.query.filter_by(username=username).first()
     try:
        db.session.delete(res)
        db.session.commit()
     except Exception as e:
         return jsonify({"message":e}), 404
     else:
         subject_name = res.subject_name
         path = bp.root_path + '/data/' + subject_name
         rmtree(path, True)
         return jsonify({"message":"删除成功" })
     
@bp.route('/upload/<subject_name>/<dirname>', methods=['POST'])
def upload_file(subject_name, dirname):
    subject_name = urllib.parse.unquote(request.headers[subject_name])
    dirname =  urllib.parse.unquote(request.headers[dirname])
    path = bp.root_path + '/data/' + subject_name + '/'+dirname   
    l = len(os.listdir(path))
    if dirname != "精彩瞬间":
        if l:
            files = os.listdir(path)
            for file in files:
                os.remove(path+'/'+file)
    file = request.files['file']
    file.save(path+'/'+file.filename)
    return jsonify({'token':"1331"})

@bp.route('/upload_mult/<subject_name>/<dirname>', methods=['POST'])
def upload_files(subject_name, dirname):
    subject_name = urllib.parse.unquote(request.headers[subject_name])
    dirname =  urllib.parse.unquote(request.headers[dirname])
    path = bp.root_path + '/data/' + subject_name + '/'+dirname
    file = request.files["file"]
   
    file.save(path+'/'+file.filename)
    return jsonify({'token':"1331"})


@bp.route('/add_subject_info', methods=['POST'])
def add_subject_info():
    form = request.form
    teacher_info = form['teacher_info']
    subject_info = form['subject_info']
    subject_name = form['subject_name']
    res = User.query.filter_by(subject_name=subject_name).first()
    res.teacher_info = teacher_info
    res.subject_info = subject_info
    tpl = DocxTemplate(bp.root_path + '/tmp.docx') 
    context ={
        "subject_name":subject_name,
        "teacher_info": teacher_info,
        "subject_info": subject_info
    }
    tpl.render(context)
    tpl.save("{}/data/{}/{}社团及老师简介.docx".format(bp.root_path,subject_name,subject_name))
    
    db.session.commit()
    return "修改成功"



@bp.route('/get_userinfos', methods=['POST'])
def get_user_infos():
    json = request.get_json()
    subject_name = json['subject_name']
    res = User.query.filter_by(subject_name=subject_name).first()
    teacher_info = res.teacher_info
    subject_info = res.subject_info
    back_url = json['back_url']
    respone = dict()
    base_path = bp.root_path + '/data/' + subject_name 
    ja = base_path + '/课时教案'
    jh= base_path + '/社团计划'
    ts =  base_path + '/特色活动方案'
    tstp = base_path + '/特色活动图片'
    jcsj = base_path + '/精彩瞬间'
    zj = base_path + '/社团总结'
    tx = base_path+'/教师风采'
    respone['base_url'] = back_url+'/'+base_path
    respone['tx'] =  os.listdir(tx)
    respone['jiaoan'] = os.listdir(ja)
    respone['jihua'] = os.listdir(jh)
    respone['subject_info'] = subject_info
    respone['teacher_info'] =teacher_info
    respone['tesefa'] = os.listdir(ts)
    respone['shunjian'] = os.listdir(jcsj)
    respone['tesetupian'] = os.listdir(tstp)
    respone['zongjie'] = os.listdir(zj)
    return jsonify(respone)


@bp.route('/remove_resource', methods=['POST'])
def remove_resource():
    data = request.get_json()
    subject_name = data['subject_name']
    dir_name =  urllib.parse.unquote(  data['dir_name'])
    image_name = data['image_name']
    file_path = "{}/data/{}/{}/{}".format(bp.root_path, subject_name, dir_name, image_name)
    os.remove(file_path)
    return ""